"""
Deterministic structured parser for pricing.xlsx — bypasses RAG entirely.
Reads the Base Pricing, Volume Tiers, and Discounts & Amendments sheets directly,
returning typed records that the comparator can use as ground truth.
"""

from __future__ import annotations
import os
import re
from dataclasses import dataclass, field
from typing import Optional
from datetime import date as Date
import openpyxl


@dataclass
class BasePriceRecord:
    sku: str
    description: str
    unit_price: float
    discount_percent: float
    tax_percent: float
    source_file: str
    sheet_name: str
    row_index: int


@dataclass
class VolumeTier:
    sku: str
    min_qty: int
    max_qty: Optional[int]  # None = no upper bound
    additional_discount: float


@dataclass
class DiscountRecord:
    sku: str
    discount_type: str  # Standard | Amendment Override | Email Addendum
    discount_percent: float
    effective_date: Optional[Date]
    expiry_date: Optional[Date]
    source_document: str  # may reference a different file (e.g. amendment_q2.docx)


@dataclass
class StructuredPricing:
    base_prices: dict[str, BasePriceRecord] = field(default_factory=dict)
    volume_tiers: dict[str, list[VolumeTier]] = field(default_factory=dict)
    discount_records: list[DiscountRecord] = field(default_factory=list)
    source_file: str = ""

    def has_data(self) -> bool:
        return bool(self.base_prices) or bool(self.volume_tiers)


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(val) -> Optional[int]:
    f = _to_float(val)
    return int(f) if f is not None else None


def _to_date(val) -> Optional[Date]:
    if val is None:
        return None
    if isinstance(val, Date):
        return val
    s = str(val).strip()
    if s.lower() in ("ongoing", "n/a", ""):
        return None
    try:
        from datetime import datetime
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_structured_pricing(file_path: str) -> StructuredPricing:
    """Walk the Excel workbook and extract typed pricing records."""
    out = StructuredPricing(source_file=os.path.basename(file_path))
    if not file_path.lower().endswith((".xlsx", ".xls")):
        return out

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return out

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        sn_lower = sheet_name.lower()
        if "base pricing" in sn_lower or sn_lower == "pricing":
            _parse_base_pricing(rows, out, sheet_name)
        elif "volume" in sn_lower or "tier" in sn_lower:
            _parse_volume_tiers(rows, out, sheet_name)
        elif "discount" in sn_lower or "amendment" in sn_lower:
            _parse_discount_log(rows, out, sheet_name)

    return out


def _parse_base_pricing(rows, out: StructuredPricing, sheet_name: str) -> None:
    """Find the header row, then read SKU pricing records."""
    header_idx = _find_header_row(rows, ["SKU"])
    if header_idx is None:
        return
    headers = [str(c or "").strip().lower() for c in rows[header_idx]]

    col_sku = _find_col(headers, ["sku"])
    col_desc = _find_col(headers, ["service description", "description", "service"])
    col_price = _find_col(headers, ["unit price", "price"])
    col_disc = _find_col(headers, ["std discount", "standard discount", "discount"])
    col_tax = _find_col(headers, ["tax rate", "tax"])

    if col_sku is None:
        return

    for r in rows[header_idx + 1:]:
        sku = str(r[col_sku] or "").strip()
        if not re.match(r"^[A-Z]{2}-\d{3}$", sku):
            continue
        unit_price = _to_float(r[col_price]) if col_price is not None else None
        discount = _to_float(r[col_disc]) if col_disc is not None else 0.0
        tax = _to_float(r[col_tax]) if col_tax is not None else 8.0
        desc = str(r[col_desc] or "").strip() if col_desc is not None else ""

        if unit_price is None:
            continue

        out.base_prices[sku] = BasePriceRecord(
            sku=sku,
            description=desc,
            unit_price=unit_price,
            discount_percent=discount or 0.0,
            tax_percent=tax or 8.0,
            source_file=out.source_file,
            sheet_name=sheet_name,
            row_index=rows.index(r) + 1,
        )


def _parse_volume_tiers(rows, out: StructuredPricing, sheet_name: str) -> None:
    """
    Volume tier sheet has multi-row headers describing volume bands like:
      'SKU' | 'Service' | '< 1,000 units' | '1,000 – 9,999 units' | '10,000 – 49,999 units' | '≥ 50,000 units'
    Each data row contains an additional discount percentage per band.
    """
    header_idx = _find_header_row(rows, ["SKU"])
    if header_idx is None:
        return

    # Find the band-defining row — search up to 4 rows ahead of header
    band_row = None
    band_row_idx = -1
    for offset in range(0, 5):
        candidate = header_idx + offset
        if candidate >= len(rows):
            break
        row = rows[candidate]
        if any(_extract_volume_band(c) for c in row if c):
            band_row = row
            band_row_idx = candidate
            break
    if band_row is None:
        return

    band_cols: dict[int, tuple[int, Optional[int]]] = {}
    for col_idx, cell in enumerate(band_row):
        band = _extract_volume_band(cell)
        if band:
            band_cols[col_idx] = band

    sku_col = _find_col([str(c or "").strip().lower() for c in rows[header_idx]], ["sku"])
    if sku_col is None:
        return

    data_start = max(header_idx, band_row_idx) + 1
    for r in rows[data_start:]:
        sku = str(r[sku_col] or "").strip() if sku_col < len(r) else ""
        if not re.match(r"^[A-Z]{2}-\d{3}$", sku):
            continue
        tiers: list[VolumeTier] = []
        for col_idx, (lo, hi) in band_cols.items():
            if col_idx >= len(r):
                continue
            disc = _to_float(r[col_idx])
            if disc is None:
                disc = 0.0
            tiers.append(
                VolumeTier(sku=sku, min_qty=lo, max_qty=hi, additional_discount=disc)
            )
        if tiers:
            out.volume_tiers[sku] = tiers


VOLUME_BAND_RE = re.compile(
    r"(?:(?P<lt><|≤|<=)\s*(?P<lt_v>[\d,]+))"
    r"|(?:(?P<gt>>|≥|>=)\s*(?P<gt_v>[\d,]+))"
    r"|(?:(?P<lo>[\d,]+)\s*[–-]\s*(?P<hi>[\d,]+))",
    re.IGNORECASE,
)


def _extract_volume_band(cell) -> Optional[tuple[int, Optional[int]]]:
    if not cell:
        return None
    text = str(cell)
    m = VOLUME_BAND_RE.search(text)
    if not m:
        return None
    if m.group("lt_v"):
        v = int(m.group("lt_v").replace(",", ""))
        return (0, v - 1)
    if m.group("gt_v"):
        v = int(m.group("gt_v").replace(",", ""))
        return (v, None)
    if m.group("lo") and m.group("hi"):
        return (
            int(m.group("lo").replace(",", "")),
            int(m.group("hi").replace(",", "")),
        )
    return None


def _parse_discount_log(rows, out: StructuredPricing, sheet_name: str) -> None:
    header_idx = _find_header_row(rows, ["SKU"])
    if header_idx is None:
        return
    headers = [str(c or "").strip().lower() for c in rows[header_idx]]

    col_sku = _find_col(headers, ["sku"])
    col_type = _find_col(headers, ["discount type", "type"])
    col_pct = _find_col(headers, ["discount (%)","discount"])
    col_eff = _find_col(headers, ["effective date", "effective"])
    col_exp = _find_col(headers, ["expiry date", "expiry", "expires"])
    col_src = _find_col(headers, ["source document", "source"])

    if col_sku is None or col_pct is None:
        return

    for r in rows[header_idx + 1:]:
        if not r or col_sku >= len(r):
            continue
        sku = str(r[col_sku] or "").strip()
        if not re.match(r"^[A-Z]{2}-\d{3}$", sku):
            continue
        pct = _to_float(r[col_pct])
        if pct is None:
            continue
        out.discount_records.append(
            DiscountRecord(
                sku=sku,
                discount_type=str(r[col_type] or "").strip() if col_type is not None else "Standard",
                discount_percent=pct,
                effective_date=_to_date(r[col_eff]) if col_eff is not None and col_eff < len(r) else None,
                expiry_date=_to_date(r[col_exp]) if col_exp is not None and col_exp < len(r) else None,
                source_document=str(r[col_src] or "").strip() if col_src is not None and col_src < len(r) else out.source_file,
            )
        )


def _find_header_row(rows, must_contain: list[str]) -> Optional[int]:
    """Find the header row where every needle is an EXACT cell match (not a substring)."""
    needles = [n.lower() for n in must_contain]
    for idx, row in enumerate(rows):
        cells = [str(c or "").strip().lower() for c in row]
        # Exact match: the cell IS the needle (or starts with it for compound headers)
        if all(any(c == n or c.startswith(n) for c in cells) for n in needles):
            return idx
    return None


def _find_col(headers: list[str], needles: list[str]) -> Optional[int]:
    """Needle-first: try each needle across all columns before moving to next needle.
    This ensures more-specific needles (earlier in list) take precedence."""
    for needle in needles:
        for idx, h in enumerate(headers):
            if h and needle in h:
                return idx
    return None


def lookup_volume_tier_discount(
    pricing: StructuredPricing, sku: str, quantity: float
) -> float:
    tiers = pricing.volume_tiers.get(sku)
    if not tiers:
        return 0.0
    qty = int(quantity)
    for tier in tiers:
        if tier.min_qty <= qty and (tier.max_qty is None or qty <= tier.max_qty):
            return tier.additional_discount
    return 0.0


def authoritative_discount(
    pricing: StructuredPricing, sku: str, invoice_date: Optional[Date]
) -> Optional[DiscountRecord]:
    """
    Walk the Discounts & Amendments log to find the active record for this SKU
    on the invoice date. Returns the record with the lowest precedence (= most authoritative).
    """
    candidates = [d for d in pricing.discount_records if d.sku == sku]
    if not candidates:
        return None

    if invoice_date:
        candidates = [
            d
            for d in candidates
            if (d.effective_date is None or d.effective_date <= invoice_date)
            and (d.expiry_date is None or d.expiry_date >= invoice_date)
        ]
    if not candidates:
        return None

    PRECEDENCE = {
        "email addendum": 1,
        "amendment override": 2,
        "amendment": 2,
        "standard": 4,
    }

    def rank(d: DiscountRecord) -> tuple[int, int]:
        prec = 4
        for key, val in PRECEDENCE.items():
            if key in d.discount_type.lower():
                prec = val
                break
        eff_ord = (
            -d.effective_date.toordinal() if d.effective_date else 0
        )  # newer effective date wins on tie
        return (prec, eff_ord)

    candidates.sort(key=rank)
    return candidates[0]
