"""Parse Excel files into sheet/row-aware chunks with metadata."""

from __future__ import annotations
import openpyxl
from app.models.schemas import Chunk


def parse_excel(file_path: str) -> list[Chunk]:
    chunks: list[Chunk] = []
    filename = file_path.split("/")[-1]
    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Build text from groups of rows (window of 5 rows, sliding by 3)
        # This keeps context while staying within token limits
        window = 5
        step = 3

        for start_idx in range(0, len(rows), step):
            end_idx = min(start_idx + window, len(rows))
            row_group = rows[start_idx:end_idx]

            lines = []
            for row in row_group:
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells)
                if any(c.strip() for c in cells):
                    lines.append(line)

            text = "\n".join(lines).strip()
            if not text:
                continue

            row_range = f"{start_idx + 1}-{end_idx}"
            chunks.append(
                Chunk(
                    text=f"[Sheet: {sheet_name}]\n{text}",
                    source_file=filename,
                    source_type="excel",
                    sheet_name=sheet_name,
                    row_range=row_range,
                    chunk_index=start_idx // step,
                    doc_precedence=3,
                )
            )

    return chunks
